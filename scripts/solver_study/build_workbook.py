"""
build_workbook.py — assembles the study's results workbook from
data/results_<version>.csv files, per the Word document's Part B §5:

  Baseline — frozen v0 rows, the "before" every later version is measured
             against. Never edited after the fact.
  Runs     — same row shape, meant to grow: append the next version's rows
             here after each experiment.
  Summary  — one row per version, computed with SUMIFS/COUNTIFS/AVERAGEIFS
             formulas against Runs (not hardcoded), so it recalculates as
             rows are appended.
  Charts   — solve-mode distribution and solve-time bars per version, read
             from Summary.

Usage:
    venv/Scripts/python.exe scripts/solver_study/build_workbook.py --versions v0
    # after a new version's results_v1.csv exists:
    venv/Scripts/python.exe scripts/solver_study/build_workbook.py --versions v0 v1
"""

import argparse
import csv
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Servings_Solver_Results.xlsx")

FONT_NAME = "Arial"
INK = "1F2620"
ACCENT = "2F5B48"
ACCENT_TEXT = "FFFFFF"
MUTED_FILL = "F5F6F2"
WARM_FILL = "FBF1DF"

HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_TEXT)
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
BODY_FONT = Font(name=FONT_NAME, size=10, color=INK)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="5B6259")
TITLE_FONT = Font(name=FONT_NAME, size=13, bold=True, color=ACCENT)
THIN = Side(style="thin", color="CBD2C7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Columns exactly as written by run_study.py's CSV_FIELDS, plus 4 helper
# abs-deviation columns appended (formulas) for Summary's AVERAGEIFS to use.
RAW_COLUMNS = [
    "version", "diet_id", "diet_label", "diet_group", "combo",
    "target_kcal", "target_protein_g", "target_carbs_g", "target_fat_g",
    "actual_kcal", "actual_protein_g", "actual_carbs_g", "actual_fat_g",
    "dev_kcal_pct", "dev_protein_pct", "dev_carbs_pct", "dev_fat_pct",
    "solve_mode", "tolerance_tier", "culinary_pass", "serving_step",
    "wall_time_ms", "lp_attempts",
    "root_cause_category", "root_cause_detail",
    "culinary_cap_adherent", "max_subrecipe_kcal_share_pct", "max_subrecipe_servings",
]
HELPER_COLUMNS = ["abs_dev_kcal_pct", "abs_dev_protein_pct", "abs_dev_carbs_pct", "abs_dev_fat_pct"]
ALL_COLUMNS = RAW_COLUMNS + HELPER_COLUMNS

HEADER_LABELS = {
    "version": "Version", "diet_id": "Diet ID", "diet_label": "Diet", "diet_group": "Group",
    "combo": "Meal Combo",
    "target_kcal": "Target Kcal", "target_protein_g": "Target Protein (g)",
    "target_carbs_g": "Target Carbs (g)", "target_fat_g": "Target Fat (g)",
    "actual_kcal": "Actual Kcal", "actual_protein_g": "Actual Protein (g)",
    "actual_carbs_g": "Actual Carbs (g)", "actual_fat_g": "Actual Fat (g)",
    "dev_kcal_pct": "Kcal Dev", "dev_protein_pct": "Protein Dev",
    "dev_carbs_pct": "Carbs Dev", "dev_fat_pct": "Fat Dev",
    "solve_mode": "Solve Mode", "tolerance_tier": "Tolerance Tier",
    "culinary_pass": "Culinary Pass", "serving_step": "Serving Step",
    "wall_time_ms": "Wall Time (ms)", "lp_attempts": "LP Attempts",
    "root_cause_category": "Root Cause", "root_cause_detail": "Root Cause Detail",
    "culinary_cap_adherent": "Culinary Guardrail OK?",
    "max_subrecipe_kcal_share_pct": "Max Subrecipe Kcal Share",
    "max_subrecipe_servings": "Max Subrecipe Servings",
    "abs_dev_kcal_pct": "|Kcal Dev|", "abs_dev_protein_pct": "|Protein Dev|",
    "abs_dev_carbs_pct": "|Carbs Dev|", "abs_dev_fat_pct": "|Fat Dev|",
}
PCT_POINT_COLUMNS = {
    "dev_kcal_pct", "dev_protein_pct", "dev_carbs_pct", "dev_fat_pct",
    "abs_dev_kcal_pct", "abs_dev_protein_pct", "abs_dev_carbs_pct", "abs_dev_fat_pct",
    "max_subrecipe_kcal_share_pct",
}

SOLVE_MODES = ["strict_tight", "strict_relaxed_tol", "relaxed_culinary", "best_effort_lp",
               "safe_fallback", "single_meal_kcal_only"]
SOLVE_MODE_LABELS = {
    "strict_tight": "Strict / 8% (Pass 1, tight)",
    "strict_relaxed_tol": "Strict / >8% (Pass 1, loose)",
    "relaxed_culinary": "Relaxed culinary (Pass 2)",
    "best_effort_lp": "Best-effort LP (Pass 3)",
    "safe_fallback": "Greedy fallback (Pass 4)",
    "single_meal_kcal_only": "Single-meal, kcal-only (v1)",
}


def load_csv(version):
    path = os.path.join(DATA_DIR, f"results_{version}.csv")
    with open(path, encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def style_header_row(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_raw_sheet(wb, name, rows, note):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value=name).font = TITLE_FONT
    ws.cell(row=2, column=1, value=note).font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ALL_COLUMNS))

    header_row = 4
    for c, col in enumerate(ALL_COLUMNS, start=1):
        ws.cell(row=header_row, column=c, value=HEADER_LABELS.get(col, col))
    style_header_row(ws, header_row, len(ALL_COLUMNS))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    dev_col_idx = {name: i + 1 for i, name in enumerate(RAW_COLUMNS)}
    for r, row in enumerate(rows, start=header_row + 1):
        for c, col in enumerate(RAW_COLUMNS, start=1):
            val = row[col]
            if col in ("target_kcal", "target_protein_g", "target_carbs_g", "target_fat_g",
                       "actual_kcal", "actual_protein_g", "actual_carbs_g", "actual_fat_g",
                       "dev_kcal_pct", "dev_protein_pct", "dev_carbs_pct", "dev_fat_pct",
                       "serving_step", "wall_time_ms", "lp_attempts",
                       "max_subrecipe_kcal_share_pct", "max_subrecipe_servings"):
                try:
                    val = float(val) if val not in ("", None) else None
                except ValueError:
                    pass
            elif col == "culinary_cap_adherent":
                val = (val == "True")
            elif col == "tolerance_tier":
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass  # "BEST_EFFORT_LP" / "SAFE_FALLBACK" stay as text
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            if col in PCT_POINT_COLUMNS:
                cell.number_format = '0.0"%"'
            elif col == "tolerance_tier" and isinstance(val, float):
                cell.number_format = "0%"

        # helper abs-deviation columns, as real formulas referencing this row
        for h, dev_col in zip(HELPER_COLUMNS, ["dev_kcal_pct", "dev_protein_pct", "dev_carbs_pct", "dev_fat_pct"]):
            src_letter = get_column_letter(dev_col_idx[dev_col])
            dst_col = ALL_COLUMNS.index(h) + 1
            cell = ws.cell(row=r, column=dst_col, value=f"=ABS({src_letter}{r})")
            cell.font = BODY_FONT
            cell.number_format = '0.0"%"'

    last_row = header_row + len(rows)
    last_col_letter = get_column_letter(len(ALL_COLUMNS))
    tbl = Table(displayName=f"{name}Data", ref=f"A{header_row}:{last_col_letter}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    widths = {"diet_label": 16, "combo": 18, "solve_mode": 20, "tolerance_tier": 12,
              "root_cause_category": 24, "root_cause_detail": 30, "culinary_cap_adherent": 14}
    for c, col in enumerate(ALL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 13)

    return header_row, last_row


def build_summary_sheet(wb, versions, runs_header_row, runs_last_row):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Summary — one row per version").font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "All cells below are formulas against the Runs sheet (SUMIFS / COUNTIFS / AVERAGEIFS / MAXIFS) — "
        "append a new version's rows to Runs, add its version id in column A here, then copy row 5's "
        "formulas down. Nothing here is a hardcoded number."
    )).font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)

    headers = (
        ["Version", "N Points"]
        + [f"% {SOLVE_MODE_LABELS[m]}" for m in SOLVE_MODES]
        + ["Avg |Kcal Dev|", "Avg |Protein Dev|", "Avg |Carbs Dev|", "Avg |Fat Dev|",
           "Avg Wall Time (ms)", "Max Wall Time (ms)", "Avg LP Attempts",
           "Guardrail Breaches", "Guardrail Breach Rate"]
    )
    header_row = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(headers))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    R = f"Runs!$A${runs_header_row + 1}:$A${runs_last_row}"  # version column
    col_letter = {col: get_column_letter(i + 1) for i, col in enumerate(ALL_COLUMNS)}

    def runs_range(col):
        letter = col_letter[col]
        return f"Runs!${letter}${runs_header_row + 1}:${letter}${runs_last_row}"

    for i, version in enumerate(versions):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=version).font = BODY_FONT
        ws.cell(row=r, column=2, value=f"=COUNTIF({R},$A{r})").font = BODY_FONT

        col = 3
        for m in SOLVE_MODES:
            f = f'=COUNTIFS({R},$A{r},{runs_range("solve_mode")},"{m}")/$B{r}'
            cell = ws.cell(row=r, column=col, value=f)
            cell.font = BODY_FONT
            cell.number_format = "0.0%"
            col += 1

        for dev_col in ["abs_dev_kcal_pct", "abs_dev_protein_pct", "abs_dev_carbs_pct", "abs_dev_fat_pct"]:
            f = f'=AVERAGEIFS({runs_range(dev_col)},{R},$A{r})'
            cell = ws.cell(row=r, column=col, value=f)
            cell.font = BODY_FONT
            cell.number_format = '0.0"%"'
            col += 1

        f = f'=AVERAGEIFS({runs_range("wall_time_ms")},{R},$A{r})'
        cell = ws.cell(row=r, column=col, value=f); cell.font = BODY_FONT; cell.number_format = "0.0"
        col += 1
        f = f'=_xlfn.MAXIFS({runs_range("wall_time_ms")},{R},$A{r})'
        cell = ws.cell(row=r, column=col, value=f); cell.font = BODY_FONT; cell.number_format = "0.0"
        col += 1
        f = f'=AVERAGEIFS({runs_range("lp_attempts")},{R},$A{r})'
        cell = ws.cell(row=r, column=col, value=f); cell.font = BODY_FONT; cell.number_format = "0.0"
        col += 1

        breach_letter = col_letter["culinary_cap_adherent"]
        breach_range = f'Runs!${breach_letter}${runs_header_row + 1}:${breach_letter}${runs_last_row}'
        f = f'=COUNTIFS({R},$A{r},{breach_range},FALSE)'
        cell = ws.cell(row=r, column=col, value=f); cell.font = BODY_FONT
        col += 1
        f = f'={get_column_letter(col-1)}{r}/$B{r}'
        cell = ws.cell(row=r, column=col, value=f); cell.font = BODY_FONT; cell.number_format = "0.0%"

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.column_dimensions["A"].width = 10

    return header_row, header_row + len(versions)


CHANGE_LOG_CONTENT = {
    "v0": {
        "date": "2026-08-11",
        "change": "Baseline — unmodified production solver, run once over the full 528-point population.",
        "result": "20.1% strict/tight, 70.3% best-effort, 0% fallback. 339/422 misses = RECIPE_POOL_MISMATCH, 77 = SERVING_BALANCE_RATIO_BLOCKED.",
        "decision": "Reference — not a candidate",
    },
    "v1": {
        "date": "2026-08-12",
        "change": "Remove max-serving ceiling (multi-meal: fully; single-meal: 8x scaled after the pickles incident); single-meal days drop all constraints/rules + macro hard bounds, kcal tolerance ladder kept.",
        "result": "Best-effort LP 70.3%->22.2% overall; real diets: best-effort 21/60->0/60, relaxed-culinary 4/60->0/60. Avg solve time 1027ms->435ms. Multi-meal unbounded-serving risk checked, worst case 22.5 servings/55% share on extreme synthetic diets only.",
        "decision": "Pending review",
    },
    "v2": {
        "date": "2026-08-12",
        "change": "Replace recipe_subrecipe_rule + flat balance ratio + max_serving with a single is_main flag per subrecipe (>=1 main/recipe, main >= every non-main, mains bounded vs. each other by MAIN_RATIO_HYPERPARAM=2.5), isolated from v1's changes.",
        "result": "Sandbox: best-effort LP 70.3%->62.5% overall. Real diets: hard-bounded solves 39/60->44/60. Zero condiment runaways in 528 points; worst case is a real dish component (Siyadiye Main) at 83.8% kcal share on an extreme synthetic diet.",
        "decision": "ACCEPTED & SHIPPED 2026-08-13 -- now the baseline",
    },
    "v2_shipped": {
        "date": "2026-08-13",
        "change": "Re-verification, not a new candidate: same 528 points run against the REAL services/mealplan_service.py + live Supabase is_main/max_serving data (verify_shipped.py). Re-run again same day after the post-cap update (see v3 row).",
        "result": "Final: 25.2% strict/tight, 65.9% best-effort. Real diets beat the sandbox's original prediction (33/60 tight vs sandbox's 24/60). Energy Bomb's zero mains checked against the full catalog and confirmed correct, not a gap -- no action needed.",
        "decision": "CONFIRMED in production",
    },
    "v3": {
        "date": "2026-08-13",
        "change": "v1's single-meal idea (skip balance + macro bounds, kcal ladder kept) reapplied to v2's shipped mechanism (is_main/MAIN_RATIO, real max_serving resolved not eliminated).",
        "result": "Real diets: best-effort LP 21/60 (v0) -> 2/60 -- biggest improvement of any version tested. Two issues found and fixed before being called done: (1) pickles-class pattern (Tarator Sauce/Pita Bread, no manual cap) -> 9 subrecipes capped directly in production; (2) single-meal path was dropping the main-vs-main ratio too (34:1 on extreme synthetic targets) -> fixed by keeping that check active regardless of skip_balance. Worst-case plate shape now matches v2 (shipped)'s own ceiling (22 servings / 55.8%). Real diets unaffected by either fix. No open limitations.",
        "decision": "Sandbox validated -- production port pending",
    },
    "v4": {
        "date": "2026-08-14",
        "change": "Proportional kcal tolerance (loosest tiers only, plateaus absolute slack for large-kcal targets) + mixed 0.25-serving step (every tolerance tier, restricted to eligible subrecipes) on top of v3. Validated on a NEW second harness: 384 real historical client days (61 real recipes, real diet targets), not just the 4-recipe synthetic population.",
        "result": "Real diets (synthetic pop.): bit-for-bit identical to v3. Real-order population (384 days) vs. shipped v2: kcal 4.55%->3.12%, roughly tied on carbs, small losses on protein/fat. Bug found and fixed before being finalized: an earlier '3 tiers not 4' simplification dropped a bounded-solution guardrail (the 20% tier), causing 7/528 synthetic + 5/384 real days to fall through to the fully unbounded best-effort fallback instead of a bounded solve -- restored the full ladder. Root single-meal-day trade-off (tight kcal, loose fat/carbs, inherited from v1) confirmed structural and traced to 53/57 of v4's regressions vs. shipped -- motivated v5.",
        "decision": "Sandbox validated -- production port pending",
    },
    "v5": {
        "date": "2026-08-14",
        "change": "Protein hard bound tried first on single-meal days (kcal + protein both bounded; carbs/fat still best-fit only), falling back cleanly to v4's fully-unbounded attempt when infeasible. Multi-meal path and final best-effort pass untouched. Requested directly in response to v4's single-meal finding.",
        "result": "Real single-meal days (110): protein 21.47%->9.87% (>2x tighter), kcal essentially unchanged (3.68%->3.96%), fat/carbs measurably looser as the direct, expected trade (25.27%->28.67%, 15.56%->17.21%). Fallback verified byte-for-byte identical to v4 on the 6/109 infeasible cases -- not limited to single-subrecipe recipes as expected, also occurs with 2-subrecipe recipes lacking protein density. Same trade-off pattern independently confirmed on the synthetic study's 10 real diets. 'Days worse than shipped' drops back to v3's level (45/384) despite the fat cost.",
        "decision": "Trade-off reviewed and ACCEPTED -- production port pending",
    },
}


def build_change_log_seed(wb, versions):
    ws = wb.create_sheet("Change Log")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Change Log").font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "Mirrors the Word document's Change Log — keep both in sync. One row per experiment."
    )).font = NOTE_FONT
    headers = ["Version", "Date", "Change & Hypothesis", "Result Summary", "Decision"]
    header_row = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(headers))
    for i, v in enumerate(versions):
        r = header_row + 1 + i
        c = CHANGE_LOG_CONTENT.get(v, {})
        decision_text = c.get("decision", "")
        shipped = "SHIPPED" in decision_text or "CONFIRMED" in decision_text
        vals = [v, c.get("date", ""), c.get("change", ""), c.get("result", ""), c.get("decision", "Pending")]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name=FONT_NAME, size=10, bold=(shipped and col in (1, 5)), color=ACCENT if shipped and col == 5 else INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [10, 12, 45, 35, 20]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_charts_sheet(wb, summary_header_row, summary_last_row, n_versions):
    ws = wb.create_sheet("Charts")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Charts").font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        f"Reads Summary!A{summary_header_row}:J{summary_last_row}. Range covers the {n_versions} version(s) "
        "present now — widen the chart's source range (right-click > Select Data) after adding new Summary rows."
    )).font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    cats = Reference(wb["Summary"], min_col=1, min_row=summary_header_row + 1, max_row=summary_last_row)

    # Column layout matches build_summary_sheet's fixed order:
    # 1=Version, 2=N Points, 3..(2+len(SOLVE_MODES))=mode %,
    # then 4 dev cols, avg time, max time, avg attempts, breaches, breach rate.
    mode_start, mode_end = 3, 2 + len(SOLVE_MODES)
    dev_start, dev_end = mode_end + 1, mode_end + 4
    time_start, time_end = dev_end + 1, dev_end + 2

    chart1 = BarChart()
    chart1.type = "col"
    chart1.grouping = "stacked"
    chart1.overlap = 100
    chart1.title = "Solve-mode distribution by version"
    chart1.y_axis.title = "% of runs"
    chart1.x_axis.title = "Version"
    data1 = Reference(wb["Summary"], min_col=mode_start, max_col=mode_end, min_row=summary_header_row, max_row=summary_last_row)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height, chart1.width = 9, 18
    ws.add_chart(chart1, "A4")

    chart2 = BarChart()
    chart2.type = "col"
    chart2.grouping = "clustered"
    chart2.title = "Avg |deviation| by macro, by version"
    chart2.y_axis.title = "Percentage points"
    chart2.x_axis.title = "Version"
    data2 = Reference(wb["Summary"], min_col=dev_start, max_col=dev_end, min_row=summary_header_row, max_row=summary_last_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height, chart2.width = 9, 18
    ws.add_chart(chart2, "A22")

    chart3 = BarChart()
    chart3.type = "col"
    chart3.title = "Avg / Max solve time by version"
    chart3.y_axis.title = "ms"
    chart3.x_axis.title = "Version"
    data3 = Reference(wb["Summary"], min_col=time_start, max_col=time_end, min_row=summary_header_row, max_row=summary_last_row)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.height, chart3.width = 9, 18
    ws.add_chart(chart3, "A40")


def build_ablation_sheet(wb):
    """Reads data/ablation_results.json (written by ablation_study.py) --
    a separate, one-off diagnostic study, not a growing version table like
    Runs, so this sheet is raw values, not SUMIFS formulas against
    anything. Re-run ablation_study.py and re-run this builder to refresh
    it after a solver_lab.py change worth re-checking."""
    path = os.path.join(DATA_DIR, "ablation_results.json")
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    summary = data["summary"]
    baseline_solved = summary[0]["solved"]
    total = summary[0]["total"]

    ws = wb.create_sheet("Ablation")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Constraint Ablation — global one-at-a-time removal").font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "From scripts/solver_study/ablation_study.py: every constraint held at its strictest (Pass 1, 8%) "
        "setting, one constraint removed globally at a time, single LP attempt per point (not the full ladder). "
        f"Baseline (nothing removed): {baseline_solved}/{total} solved — matches v0's strict_tight count exactly. "
        "Raw values, not formulas — re-run ablation_study.py then this builder to refresh."
    )).font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ["Constraint removed", "Solved", "Total", "% Solved", "Points Unlocked", "% of Population"]
    header_row = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(headers))

    ranked = sorted(summary[1:], key=lambda s: -(s["solved"] - baseline_solved))
    ordered = [summary[0]] + ranked
    for i, s in enumerate(ordered):
        r = header_row + 1 + i
        delta = s["solved"] - baseline_solved
        vals = [s["label"], s["solved"], s["total"], s["solved"] / s["total"], delta, delta / total]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            if c in (4, 6):
                cell.number_format = "0.0%"
        if i == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10, italic=True, color="5B6259")

    last_row = header_row + len(ordered)
    widths = [40, 10, 10, 12, 16, 16]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    chart = BarChart()
    chart.type = "bar"  # horizontal, reads better for a ranked list
    chart.title = "Points unlocked by removing one constraint (of 528)"
    chart.x_axis.title = "Constraint removed"
    chart.y_axis.title = "Points unlocked"
    data_ref = Reference(ws, min_col=5, min_row=header_row + 2, max_row=last_row)  # skip baseline row (delta=0)
    cats_ref = Reference(ws, min_col=1, min_row=header_row + 2, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    chart.height, chart.width = 12, 20
    ws.add_chart(chart, "A" + str(last_row + 3))

    return header_row, last_row


MACRO_FREE_COLUMNS = [
    "diet_id", "diet_label", "diet_group", "combo",
    "target_kcal", "target_protein_g", "target_carbs_g", "target_fat_g",
    "actual_kcal", "actual_protein_g", "actual_carbs_g", "actual_fat_g",
    "dev_kcal_pct", "dev_protein_pct", "dev_carbs_pct", "dev_fat_pct",
    "tier", "culinary_pass", "serving_step",
]
MACRO_FREE_LABELS = {
    "diet_id": "Diet ID", "diet_label": "Diet", "diet_group": "Group", "combo": "Meal Combo",
    "target_kcal": "Target Kcal", "target_protein_g": "Target Protein (g)",
    "target_carbs_g": "Target Carbs (g)", "target_fat_g": "Target Fat (g)",
    "actual_kcal": "Actual Kcal", "actual_protein_g": "Actual Protein (g)",
    "actual_carbs_g": "Actual Carbs (g)", "actual_fat_g": "Actual Fat (g)",
    "dev_kcal_pct": "Kcal Dev", "dev_protein_pct": "Protein Dev",
    "dev_carbs_pct": "Carbs Dev", "dev_fat_pct": "Fat Dev",
    "tier": "Tier", "culinary_pass": "Culinary Pass", "serving_step": "Serving Step",
}
MACRO_FREE_PCT_COLS = {"dev_kcal_pct", "dev_protein_pct", "dev_carbs_pct", "dev_fat_pct"}


def build_macro_free_sheet(wb):
    """Reads data/macro_free_deviation.json (written by
    macro_free_deviation_study.py) -- another one-off diagnostic, raw
    values not formulas, same rationale as build_ablation_sheet. Answers a
    different question than Ablation: not "does it become feasible" but
    "once macro bands are removed, how close does the best fit actually
    land." No v1 exists yet from this -- purely feeds the decision on what
    a real macro-tolerance change should be sized to."""
    path = os.path.join(DATA_DIR, "macro_free_deviation.json")
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    rows = data["rows"]
    solved = [r for r in rows if r["tier"] != "INFEASIBLE"]

    ws = wb.create_sheet("Macro-Free Fit")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Macro-Free Best-Fit Deviation").font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "From scripts/solver_study/macro_free_deviation_study.py: protein/carbs/fat hard bounds removed, "
        "kcal held at strict/8%, existing objective finds its best fit. No v1 exists yet -- diagnostic only. "
        "Raw values, not formulas — re-run the script then this builder to refresh."
    )).font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(MACRO_FREE_COLUMNS))

    header_row = 4
    for c, col in enumerate(MACRO_FREE_COLUMNS, start=1):
        ws.cell(row=header_row, column=c, value=MACRO_FREE_LABELS.get(col, col))
    style_header_row(ws, header_row, len(MACRO_FREE_COLUMNS))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for r, row in enumerate(rows, start=header_row + 1):
        for c, col in enumerate(MACRO_FREE_COLUMNS, start=1):
            val = row.get(col)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            if col in MACRO_FREE_PCT_COLS and val is not None:
                cell.number_format = '0.0"%"'

    last_row = header_row + len(rows)
    last_col_letter = get_column_letter(len(MACRO_FREE_COLUMNS))
    tbl = Table(displayName="MacroFreeData", ref=f"A{header_row}:{last_col_letter}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    widths = {"diet_label": 16, "combo": 18, "tier": 22, "culinary_pass": 12}
    for c, col in enumerate(MACRO_FREE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 13)

    # ---- Summary block, off to the right so it's visible without
    # scrolling past 528 raw rows -----------------------------------------
    S = len(MACRO_FREE_COLUMNS) + 2  # start column, 1 gap after raw table
    ws.cell(row=4, column=S, value="Summary — avg |deviation| by group").font = TITLE_FONT

    def abs_stats(group_rows, key):
        vals = sorted(abs(r[key]) for r in group_rows if r.get(key) is not None)
        n = len(vals)
        if n == 0:
            return 0, 0, 0, 0
        return sum(vals) / n, vals[n // 2], vals[int(n * 0.9)], vals[-1]

    r0 = 6
    headers = ["Group", "Protein avg/median", "Carbs avg/median", "Fat avg/median", "Kcal avg/median"]
    for c, h in enumerate(headers, start=0):
        ws.cell(row=r0, column=S + c, value=h)
    style_header_row(ws, r0, len(headers))
    for i, group in enumerate(["real", "synthetic"]):
        sub = [r for r in solved if r["diet_group"] == group]
        r = r0 + 1 + i
        ws.cell(row=r, column=S, value=group.capitalize()).font = BODY_FONT
        for c, key in enumerate(["dev_protein_pct", "dev_carbs_pct", "dev_fat_pct", "dev_kcal_pct"], start=1):
            avg, med, _p90, _mx = abs_stats(sub, key)
            cell = ws.cell(row=r, column=S + c, value=f"{avg:.1f}% / {med:.1f}%")
            cell.font = BODY_FONT

    r1 = r0 + 5
    ws.cell(row=r1, column=S, value="Real diets, avg |deviation| by combo").font = TITLE_FONT
    headers2 = ["Combo", "Protein", "Carbs", "Fat", "Kcal"]
    hr2 = r1 + 2
    for c, h in enumerate(headers2, start=0):
        ws.cell(row=hr2, column=S + c, value=h)
    style_header_row(ws, hr2, len(headers2))

    real_solved = [r for r in solved if r["diet_group"] == "real"]
    combos = list(dict.fromkeys(r["combo"] for r in real_solved))
    combo_row0 = hr2 + 1
    for i, combo in enumerate(combos):
        sub = [r for r in real_solved if r["combo"] == combo]
        r = combo_row0 + i
        ws.cell(row=r, column=S, value=combo).font = BODY_FONT
        for c, key in enumerate(["dev_protein_pct", "dev_carbs_pct", "dev_fat_pct", "dev_kcal_pct"], start=1):
            avg, _med, _p90, _mx = abs_stats(sub, key)
            cell = ws.cell(row=r, column=S + c, value=round(avg, 1))
            cell.font = BODY_FONT
            cell.number_format = '0.0"%"'
    combo_last_row = combo_row0 + len(combos) - 1

    for c in range(S, S + 5):
        ws.column_dimensions[get_column_letter(c)].width = 20

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Real diets — avg |deviation| by macro, by meal combo"
    chart.y_axis.title = "Percentage points"
    chart.x_axis.title = "Meal combo"
    data_ref = Reference(ws, min_col=S + 1, max_col=S + 4, min_row=hr2, max_row=combo_last_row)
    cats_ref = Reference(ws, min_col=S, min_row=combo_row0, max_row=combo_last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height, chart.width = 10, 20
    ws.add_chart(chart, get_column_letter(S) + str(combo_last_row + 3))

    return header_row, last_row


REAL_ORDER_COLUMNS = [
    "version", "day_id", "date", "meal_types",
    "target_kcal", "target_protein_g", "target_carbs_g", "target_fat_g",
    "actual_kcal", "actual_protein_g", "actual_carbs_g", "actual_fat_g",
    "dev_kcal_pct", "dev_protein_pct", "dev_carbs_pct", "dev_fat_pct",
    "solve_mode", "tolerance_tier", "culinary_pass", "serving_step",
    "wall_time_ms", "lp_attempts", "culinary_cap_adherent",
]
REAL_ORDER_LABELS = {
    "version": "Version", "day_id": "Day ID", "date": "Date", "meal_types": "Meal Types",
    "target_kcal": "Target Kcal", "target_protein_g": "Target Protein (g)",
    "target_carbs_g": "Target Carbs (g)", "target_fat_g": "Target Fat (g)",
    "actual_kcal": "Actual Kcal", "actual_protein_g": "Actual Protein (g)",
    "actual_carbs_g": "Actual Carbs (g)", "actual_fat_g": "Actual Fat (g)",
    "dev_kcal_pct": "Kcal Dev", "dev_protein_pct": "Protein Dev",
    "dev_carbs_pct": "Carbs Dev", "dev_fat_pct": "Fat Dev",
    "solve_mode": "Solve Mode", "tolerance_tier": "Tolerance Tier",
    "culinary_pass": "Culinary Pass", "serving_step": "Serving Step",
    "wall_time_ms": "Wall Time (ms)", "lp_attempts": "LP Attempts",
    "culinary_cap_adherent": "Culinary Guardrail OK?",
}
REAL_ORDER_PCT_COLS = {"dev_kcal_pct", "dev_protein_pct", "dev_carbs_pct", "dev_fat_pct"}
REAL_ORDER_VERSIONS = ["v2_shipped", "v3", "v4", "v5"]


def _load_real_order_rows():
    """Merges real_orders_results.json (v3/v4/v5) and
    real_orders_results_shipped.json (v2_shipped) into one flat row list,
    normalized to REAL_ORDER_COLUMNS. Returns None if the data isn't
    present (this study's real-order harness hasn't been run yet)."""
    main_path = os.path.join(DATA_DIR, "real_orders_results.json")
    shipped_path = os.path.join(DATA_DIR, "real_orders_results_shipped.json")
    if not (os.path.exists(main_path) and os.path.exists(shipped_path)):
        return None

    with open(main_path, encoding="utf-8") as f:
        main_results = json.load(f)["results"]
    with open(shipped_path, encoding="utf-8") as f:
        shipped_results = json.load(f)["results"]

    rows = []
    for r in shipped_results + main_results:
        rows.append({
            "version": r["version"], "day_id": r["day_id"], "date": r["date"],
            "meal_types": "|".join(r["meal_types"]),
            "target_kcal": r["target"]["kcal"], "target_protein_g": r["target"]["protein_g"],
            "target_carbs_g": r["target"]["carbs_g"], "target_fat_g": r["target"]["fat_g"],
            "actual_kcal": r["actual"]["kcal"], "actual_protein_g": r["actual"]["protein_g"],
            "actual_carbs_g": r["actual"]["carbs_g"], "actual_fat_g": r["actual"]["fat_g"],
            "dev_kcal_pct": r["deviation_pct"]["kcal"], "dev_protein_pct": r["deviation_pct"]["protein"],
            "dev_carbs_pct": r["deviation_pct"]["carbs"], "dev_fat_pct": r["deviation_pct"]["fat"],
            "solve_mode": r["solve_mode"], "tolerance_tier": r["tolerance_tier"],
            "culinary_pass": r["culinary_pass"], "serving_step": r["serving_step"],
            "wall_time_ms": r["wall_time_ms"], "lp_attempts": r["lp_attempts"],
            "culinary_cap_adherent": r["culinary_cap_adherent"],
        })
    return rows


def build_real_orders_sheet(wb):
    """Reads real_orders_results.json + real_orders_results_shipped.json --
    the second validation harness (see Word doc §16): 384 REAL historical
    client days (61 real recipes, real diet targets pulled from
    daily_macro_target), re-solved through v3/v4/v5, plus the actual
    current shipped services/mealplan_service.py (v2_shipped) on the same
    384 days for a controlled comparison. One-off study, not part of the
    growing Runs/Summary system (different population shape) -- raw
    values, same rationale as build_ablation_sheet. Re-run
    pull_real_orders.py + run_real_orders.py + run_real_orders_shipped.py,
    then this builder, to refresh."""
    rows = _load_real_order_rows()
    if rows is None:
        return None

    versions_present = [v for v in REAL_ORDER_VERSIONS if any(r["version"] == v for r in rows)]

    ws = wb.create_sheet("Real-Order Validation")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Real-Order Validation — 384 Real Historical Client Days").font = TITLE_FONT
    ws.cell(row=2, column=1, value=(
        "Second validation harness (Word doc §16): real recipes/targets pulled live from Supabase, not the "
        "4-recipe synthetic population. 'v2_shipped' re-runs the ACTUAL current services/mealplan_service.py "
        "on the same 384 days for a controlled comparison -- not the historical delivered-servings record, "
        "which turned out to be confounded (see §16). Raw values, not formulas -- re-run the pull + solve "
        "scripts then this builder to refresh."
    )).font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(REAL_ORDER_COLUMNS))

    # ---- Summary block first, so it's visible without scrolling past 1500+ raw rows ----
    ws.cell(row=4, column=1, value="Summary — avg |deviation|, wall time, attempts by version").font = TITLE_FONT
    headers = ["Version", "N Days", "Avg |Kcal Dev|", "Avg |Protein Dev|", "Avg |Carbs Dev|", "Avg |Fat Dev|",
               "Avg Wall Time (ms)", "Avg LP Attempts", "Guardrail Breaches"]
    hr0 = 6
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr0, column=c, value=h)
    style_header_row(ws, hr0, len(headers))

    for i, version in enumerate(versions_present):
        sub = [r for r in rows if r["version"] == version]
        n = len(sub)
        r = hr0 + 1 + i
        ws.cell(row=r, column=1, value=version).font = BODY_FONT
        ws.cell(row=r, column=2, value=n).font = BODY_FONT
        for c, dev_col in enumerate(["dev_kcal_pct", "dev_protein_pct", "dev_carbs_pct", "dev_fat_pct"], start=3):
            avg = sum(abs(x[dev_col]) for x in sub) / n
            cell = ws.cell(row=r, column=c, value=round(avg, 2))
            cell.font = BODY_FONT
            cell.number_format = '0.00"%"'
        avg_wall = sum(x["wall_time_ms"] for x in sub) / n
        cell = ws.cell(row=r, column=7, value=round(avg_wall, 1)); cell.font = BODY_FONT
        avg_attempts = sum(x["lp_attempts"] for x in sub) / n
        cell = ws.cell(row=r, column=8, value=round(avg_attempts, 2)); cell.font = BODY_FONT
        breaches = sum(1 for x in sub if not x["culinary_cap_adherent"])
        cell = ws.cell(row=r, column=9, value=breaches); cell.font = BODY_FONT

    summary_last_row = hr0 + len(versions_present)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 17

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Real-order population — avg |deviation| by macro, by version"
    chart.y_axis.title = "Percentage points"
    chart.x_axis.title = "Version"
    data_ref = Reference(ws, min_col=3, max_col=6, min_row=hr0, max_row=summary_last_row)
    cats_ref = Reference(ws, min_col=1, min_row=hr0 + 1, max_row=summary_last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height, chart.width = 10, 20
    ws.add_chart(chart, "A" + str(summary_last_row + 3))

    # ---- Raw per-day table, all versions stacked ----
    raw_start = summary_last_row + 24
    ws.cell(row=raw_start, column=1, value="Raw data — one row per (version, day)").font = TITLE_FONT
    header_row = raw_start + 2
    for c, col in enumerate(REAL_ORDER_COLUMNS, start=1):
        ws.cell(row=header_row, column=c, value=REAL_ORDER_LABELS.get(col, col))
    style_header_row(ws, header_row, len(REAL_ORDER_COLUMNS))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for r, row in enumerate(rows, start=header_row + 1):
        for c, col in enumerate(REAL_ORDER_COLUMNS, start=1):
            val = row.get(col)
            if col == "culinary_cap_adherent":
                val = bool(val)
            elif col == "tolerance_tier":
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            if col in REAL_ORDER_PCT_COLS and isinstance(val, (int, float)):
                cell.number_format = '0.0"%"'
            elif col == "tolerance_tier" and isinstance(val, float):
                cell.number_format = "0%"

    last_row = header_row + len(rows)
    last_col_letter = get_column_letter(len(REAL_ORDER_COLUMNS))
    tbl = Table(displayName="RealOrderData", ref=f"A{header_row}:{last_col_letter}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    widths = {"date": 12, "meal_types": 22, "solve_mode": 20, "tolerance_tier": 12, "culinary_cap_adherent": 14}
    for c, col in enumerate(REAL_ORDER_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 13)

    return header_row, last_row


def build_cover_sheet(wb, versions, has_ablation, has_macro_free, has_real_orders):
    ws = wb.create_sheet("Read Me", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    STATUS_FONT = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT)
    lines = [
        ("Servings Solver — Results Workbook", TITLE_FONT),
        ("", BODY_FONT),
        ("STATUS (2026-08-14): v2 (the is_main mechanism) has SHIPPED to production. v3, v4, v5 are all", STATUS_FONT),
        ("sandbox-validated with no open limitations, none yet ported. v4/v5 also validated on a second,", STATUS_FONT),
        ("real-order harness -- see the Real-Order Validation sheet and the Word doc's §16-18.", STATUS_FONT),
        ("", BODY_FONT),
        ("Companion to Servings_Solver_Improvement_Study.docx (Part B). See that document for the full", NOTE_FONT),
        ("methodology, the diet population, the fixed recipe combination, and the decision protocol.", NOTE_FONT),
        ("", BODY_FONT),
        ("Baseline   — frozen v0 rows (today's unmodified solver). Reference only, never edited.", BODY_FONT),
        ("Runs       — grows over time: append each new version's rows here after an experiment.", BODY_FONT),
        ("Summary    — one row per version, all formulas against Runs. Copy the last row's formulas", BODY_FONT),
        ("             down and set the new version id in column A after appending to Runs.", BODY_FONT),
        ("Charts     — solve-mode / deviation / time comparisons across versions, reads Summary.", BODY_FONT),
        ("Change Log — mirrors the Word document's Change Log.", BODY_FONT),
    ]
    if has_ablation:
        lines.append(("Ablation   — one-off: which single constraint, removed globally, unlocks the most points.", BODY_FONT))
    if has_macro_free:
        lines.append(("Macro-Free Fit — one-off: with macro bands removed, how close does the best fit land?", BODY_FONT))
    if has_real_orders:
        lines.append(("Real-Order Validation — one-off: 384 REAL historical client days (61 real recipes,", BODY_FONT))
        lines.append(("             real diet targets), v2_shipped/v3/v4/v5 compared head-to-head. Doc §16-18.", BODY_FONT))
    lines.append(("", BODY_FONT))
    lines.append((f"Versions currently in this workbook (synthetic population): {', '.join(versions)}", BODY_FONT))
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--versions", nargs="+", default=["v0"])
    args = ap.parse_args()

    all_rows = []
    for v in args.versions:
        all_rows.extend(load_csv(v))
    baseline_rows = [r for r in all_rows if r["version"] == "v0"]

    ablation_path = os.path.join(DATA_DIR, "ablation_results.json")
    has_ablation = os.path.exists(ablation_path)
    macro_free_path = os.path.join(DATA_DIR, "macro_free_deviation.json")
    has_macro_free = os.path.exists(macro_free_path)
    has_real_orders = os.path.exists(os.path.join(DATA_DIR, "real_orders_results.json")) and \
        os.path.exists(os.path.join(DATA_DIR, "real_orders_results_shipped.json"))

    wb = Workbook()
    wb.remove(wb.active)

    build_cover_sheet(wb, args.versions, has_ablation, has_macro_free, has_real_orders)
    write_raw_sheet(wb, "Baseline", baseline_rows,
                     "Frozen v0 — the unmodified production solver's result on every (diet, combo) point. Never edited.")
    runs_header_row, runs_last_row = write_raw_sheet(
        wb, "Runs", all_rows,
        "Grows over time — append the next version's rows here after each sandbox experiment."
    )
    summary_header_row, summary_last_row = build_summary_sheet(wb, args.versions, runs_header_row, runs_last_row)
    build_change_log_seed(wb, args.versions)
    build_charts_sheet(wb, summary_header_row, summary_last_row, len(args.versions))
    if has_ablation:
        build_ablation_sheet(wb)
    if has_macro_free:
        build_macro_free_sheet(wb)
    if has_real_orders:
        build_real_orders_sheet(wb)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  Baseline: {len(baseline_rows)} rows")
    print(f"  Runs: {len(all_rows)} rows ({len(args.versions)} version(s): {', '.join(args.versions)})")
    if has_ablation:
        print("  Ablation: included")
    if has_macro_free:
        print("  Macro-Free Fit: included")
    if has_real_orders:
        print("  Real-Order Validation: included")


if __name__ == "__main__":
    main()
